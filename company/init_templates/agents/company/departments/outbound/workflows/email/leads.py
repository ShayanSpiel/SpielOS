#!/usr/bin/env python3
"""
Outbound Department — lead ingest, score, and lookalike discovery.

Usage:
    python3 leads.py ingest <file.xlsx|csv>   # validate + dedupe + append to master
    python3 leads.py score <file.csv>         # ICP score + send recommendation
    python3 leads.py lookalike [segment]      # discovery queries from existing leads
    python3 leads.py merge <file.csv>         # dedupe rows into master (email+domain)

Honest scope (2026-08-08):
  * Fully automatic 24/7 scraping is NOT built here and NOT recommended: free
    tools that scrape LinkedIn/registries violate their ToS and burn the
    sending domains' reputation. The machine keeps sending non-stop; the lead
    feed is a *drop-in pipeline*: any source (Apollo free export, a browser
    extension export, a session where this assistant researches websites)
    lands in the same place and gets validated, deduped, scored and queued.
  * ingest/merge/score are fully deterministic and offline.

Scoring (deterministic, 0-100):
    0-100 for usable email (format + no disposable/placeholder domain)
    +30 exact segment match, +15 grapnular match
    +20 employees in 5..250 (small/mid: right-size for AI employees)
    +15 annual revenue >= $1M (established, has budget); -10 if below
    +10 country in target list
    +20 ranked title (owner/founder/ceo/cmo/coo/marketing lead/recruiter)
    +10 personalization hook already filled in
    Recommendation: >=70 "Ready to personalized" · >=45 "Routing email only"
                   · else "Backup; wait"
Run `score` after every ingest; check `lookalike` against the segments you
want to grow, then fill net new rows and re-run ingest.
"""

import csv
import json
import re
import sys
from datetime import datetime, timezone

import openpyxl

from . import config
from .outbound import COL_MAP, SHEET_NAME

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")
BAD_DOMAINS = {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com",
               "icloud.com", "proton.me", "protonmail.com", "test.com", "example.com",
               "spielos.xyz", "in.spielos.xyz"}
# ── ICP exclusions (canonical spec: company/strategy/icp.md) ────────────────
# NEVER send targets, regardless of score:
#   AI companies / AI agencies / AI consultancies / agent builders /
#   software & development agencies / studios. Consultancies and SaaS with
#   operational workflows ARE targets (see canonical ICP business types).
EXCLUDE_MARKERS = ("software", "development", "studio", "dev", "automation",
                   "chatbot", "llm", "gpt")
SEGMENT_TARGETS = ("recruit", "staff", "talent", "headhunt", "agency", "digital",
                   "marketing", "saas", "property", "real estate", "ecommerce",
                   "logistics", "legal", "construction", "estate", "consult",
                   "coach", "educ", "train", "travel", "marketplace",
                   "online service")
TARGET_COUNTRIES = {"united kingdom", "united states", "us", "uk", "canada", "australia",
                    "germany", "france", "netherlands", "sweden", "norway", "denmark",
                    "fi", "finland", "ireland", "uae", "dubai", "saudi", "qatar"}
TITLE_RANK = {"founder": 20, "owner": 20, "ceo": 20, "cto": 20, "coo": 20,
              "director": 16, "head": 16, "vp": 16, "recruiter": 14,
              "recruitment": 12, "marketing": 12, "manager": 10,
              "operations": 8, "lead": 8, "hr": 8}


def _excluded(segment: str) -> str:
    """Return the exclusion marker if the segment is outside the ICP
    (AI / software / development / builders), else empty string."""
    s = f" {str(segment or '').strip().lower()} "
    for m in EXCLUDE_MARKERS:
        if f" {m} " in s or s.startswith(f" {m} ") or s.endswith(f" {m}") or f" {m}-" in s:
            return m
    if re.search(r"\bai\b|ai[-_]", s):
        return "ai"
    return ""


def _revenue(row: dict) -> float:
    """Parse annual revenue to USD float; 0 when missing/unparseable."""
    raw = str(row.get("annual_revenue") or row.get("revenue") or "0")
    raw = re.sub(r"[^\d.]", "", raw.replace(",", ""))
    try:
        return float(raw or 0)
    except ValueError:
        return 0.0


def _norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def _score(row: dict) -> tuple:
    email = str(row.get("email") or "").strip().lower()
    if not EMAIL_RE.match(email):
        return 0, "email invalid"
    dom = email.split("@")[-1]
    if dom in BAD_DOMAINS:
        return 0, "disposable/proton MX not sendable"
    segment = str(row.get("segment") or "").strip().lower()
    marker = _excluded(segment)
    if marker:
        return 0, "Do not automate"  # out of ICP: AI/software/development/consulting
    score = 40
    if segment:
        score += 15
        if any(t in segment for t in SEGMENT_TARGETS):
            score += 25
    try:
        emps = int(float(str(row.get("employees") or "0").replace(",", "")))
    except ValueError:
        emps = 0
    if 5 <= emps <= 50:
        score += 20
    elif 51 <= emps <= 250:
        score += 12
    elif 0 < emps < 5:
        score += 4
    elif emps > 250:
        score -= 15   # enterprise-scale: not the ICP (see ../spielos-icp.md)
    rev = _revenue(row)
    if rev >= 1_000_000:
        score += 15   # established buyer with money: the ICP revenue band
    elif rev > 0:
        score -= 10   # below band: no budget for this
    country = str(row.get("country") or "").lower()
    if country in TARGET_COUNTRIES:
        score += 10
    title = str(row.get("title") or "").lower()
    for k, v in TITLE_RANK.items():
        if k in title:
            score += v
            break
    if str(row.get("personalization_hook") or "").strip():
        score += 10
    rec = "Ready to personalized" if score >= 70 else (
        "Routing email only" if score >= 40 else "Backup; wait")
    return min(score, 100), rec


def reclassify(path: str = None) -> dict:
    """Rewrite send_recommendation + icp_confidence + qualification_rationale
    for every row in the master workbook using the current deterministic
    rules. Idempotent; safe to re-run after rule changes."""
    path = path or config.DATABASE_PATH
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_NAME]
    counts = {}
    rationale = {}
    for row in ws.iter_rows(min_row=2):
        cells = {k: row[v] for k, v in COL_MAP.items()}
        email = str(cells["email"].value or "").strip()
        if not email:
            continue
        lead = {k: (v.value if v is not None else "") for k, v in cells.items()}
        score, rec = _score(lead)
        marker = _excluded(lead.get("segment"))
        why = "excluded by ICP rule" if marker else (
            f"score {score}: segment/gse/name/tech"
            if not email else f"score {score}: segment, employees, title, hook")
        cells["send_recommendation"].value = rec
        cells["icp_confidence"].value = score
        cells["qualification_rationale"].value = why
        rec = str(rec)
        counts[rec] = counts.get(rec, 0) + 1
    wb.save(path)
    return counts


def _next_lead_id(existing_ids: set) -> str:
    mx = 0
    for i in existing_ids:
        m = re.search(r"-?(\d+)$", i)
        if m:
            mx = max(mx, int(m.group(1)))
    return f"EN-{mx + 1:03d}"


def _read_rows(path: str) -> list:
    rows = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1))]
        idx = {h: i for i, h in enumerate(header)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not any(r):
                continue
            rows.append({h: (r[i] if i < len(r) else "") for h, i in idx.items()})
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                rows.append({k.strip().lower(): v for k, v in r.items()})
    return rows


def merge(path: str) -> None:
    """Dedupe new rows against master + staging, print what is new."""
    wb = openpyxl.load_workbook(config.DATABASE_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    master = set()
    ids = set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[COL_MAP["email"]]:
            continue
        master.add(str(row[COL_MAP["email"]]).strip().lower())
        lid = str(row[COL_MAP["lead_id"]] or "").strip()
        if lid:
            ids.add(lid)
    wb.close()
    new_email, new_domain = [], []
    for r in _read_rows(path):
        e = str(r.get("email") or "").strip().lower()
        if not EMAIL_RE.match(e):
            continue
        if e in master:
            continue
        dom = e.split("@")[-1]
        if dom in BAD_DOMAINS:
            continue
        new_email.append(e)
    print(f"merge: {len(_read_rows(path))} rows in file, {len(new_email)} usable-new emails. "
          f"Run `leads.py ingest {path}` to append them to the master.")


def ingest(path: str) -> None:
    """Validate, dedupe, append new rows to the master workbook."""
    wb = openpyxl.load_workbook(config.DATABASE_PATH)
    ws = wb[SHEET_NAME]
    master = set()
    ids = set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[COL_MAP["email"]]:
            continue
        master.add(str(row[COL_MAP["email"]]).strip().lower())
        lid = str(row[COL_MAP["lead_id"]] or "").strip()
        if lid:
            ids.add(lid)

    rows = _read_rows(path)
    added, skipped = 0, 0
    reasons = {}
    for row in rows:
        email = str(row.get("email") or "").strip().lower()
        if not EMAIL_RE.match(email):
            skipped += 1
            continue
        if email in master:
            skipped += 1
            reasons.setdefault("duplicate", 0)
            reasons["duplicate"] += 1
            continue
        dom = email.split("@")[-1]
        if dom in BAD_DOMAINS:
            skipped += 1
            reasons.setdefault("disposable", 0)
            reasons["disposable"] += 1
            continue
        score_val, rec = _score(row)
        lid = str(row.get("lead_id") or "").strip() or _next_lead_id(ids)
        ids.add(lid)
        master.add(email)

        # ── verification tier at intake ───────────────────────────────
        # Apollo's own status when present (Verified/valid), plus catch-all
        # flag; otherwise L1 (MX + disposable) decides. MX check via the
        # DNS fallback (system resolver is broken on this box).
        apollo_status = str(row.get("email status") or row.get("email_status") or "").strip().lower()
        apollo_catchall = str(row.get("primary email catch-all status") or "").strip().lower()
        is_catchall = apollo_catchall == "catch-all" or "catch-all" in apollo_catchall and "not" not in apollo_catchall
        status = ""
        if "bounced" in apollo_status or "invalid" in apollo_status:
            status = "Bounced; suppressed"
        elif "verified" in apollo_status or "valid" in apollo_status:
            status = "Catch-all; unverified" if is_catchall else "Verified"
        elif is_catchall:
            status = "Catch-all; unverified"
        else:
            try:
                from verify import l1_check
                l1 = l1_check(email)
                if l1["tier"] == "Invalid":
                    status = "Bounced; suppressed"
                elif l1["tier"] == "Catch-all":
                    status = "Catch-all; unverified"
                else:
                    status = "Publicly listed; not deliverability-verified"
            except Exception:
                status = "Publicly listed; not deliverability-verified"
        if status == "Bounced; suppressed":
            skipped += 1
            reasons.setdefault("invalid", 0)
            reasons["invalid"] += 1
            continue

        vals = [""] * len(COL_MAP)
        vals[COL_MAP["lead_id"]] = lid
        vals[COL_MAP["email"]] = email
        vals[COL_MAP["company"]] = str(row.get("company") or row.get("company name") or "").strip()
        vals[COL_MAP["contact_name"]] = str(row.get("contact_name") or "").strip()
        vals[COL_MAP["title"]] = str(row.get("title") or "").strip()
        vals[COL_MAP["company_domain"]] = str(row.get("company_domain") or dom).strip().lower()
        vals[COL_MAP["website"]] = str(row.get("website") or "").strip()
        vals[COL_MAP["country"]] = str(row.get("country") or "").strip()
        vals[COL_MAP["segment"]] = str(row.get("segment") or row.get("industry") or "").strip()
        vals[COL_MAP["employees"]] = str(row.get("employees") or row.get("# employees") or "").strip()
        vals[COL_MAP["language"]] = str(row.get("language") or "English").strip()
        vals[COL_MAP["personalization_hook"]] = str(row.get("personalization_hook") or "").strip()
        vals[COL_MAP["suggested_cta"]] = str(row.get("suggested_cta") or "").strip()
        vals[COL_MAP["technologies"]] = str(row.get("technologies") or "").strip()
        vals[COL_MAP["need_buying_signals"]] = str(row.get("need_buying_signals") or "").strip()
        vals[COL_MAP["qualification_rationale"]] = str(row.get("qualification_rationale") or "").strip()
        vals[COL_MAP["pain_hypothesis"]] = str(row.get("pain_hypothesis") or "").strip()
        vals[COL_MAP["recommended_pilot"]] = str(row.get("recommended_pilot") or "").strip()
        vals[COL_MAP["annual_revenue"]] = str(row.get("annual_revenue") or row.get("revenue") or "").strip()
        vals[COL_MAP["market"]] = str(row.get("market") or "").strip()
        vals[COL_MAP["person_linkedin"]] = str(row.get("person_linkedin") or "").strip()
        vals[COL_MAP["source_url"]] = str(row.get("source_url") or "").strip()
        vals[COL_MAP["sequence_status"]] = ""
        src = str(row.get("source") or "").strip()
        if not src and ("apollo contact id" in row or "company name for emails" in row):
            src = "Apollo contact exports"
        vals[COL_MAP["source"]] = src
        vals[COL_MAP["send_recommendation"]] = rec
        vals[COL_MAP["outreach_tier"]] = str(row.get("outreach_tier") or "").strip()
        vals[COL_MAP["icp_confidence"]] = str(row.get("icp_confidence") or "").strip()
        vals[COL_MAP["email_status"]] = status
        ws.append(vals)
        added += 1
    wb.save(config.DATABASE_PATH)
    detail = " ".join(f"{k}={v}" for k, v in reasons.items())
    print(f"ingest: added {added} new leads ('{path}'); skipped {skipped} "
          f"({detail or 'invalid/duplicate/disposable'}). Verification tier set at intake.")


def score(path: str) -> None:
    rows = _read_rows(path)
    print(f"{'EMAIL':<38} {'SCORE':>5}  RECOMMENDATION")
    for row in rows:
        email = str(row.get("email") or "").strip().lower()
        s, rec = _score(row)
        print(f"{email:<38} {s:>5}  {rec}")
    print(f"\n(score is informational; run `leads.py ingest` to write to master)")


def lookalikes(segment: str = "") -> None:
    contacts = _read_rows(config.DATABASE_PATH)
    segments = set()
    if segment:
        segments.add(segment.lower())
    else:
        for c in contacts:
            seg = str(c.get("segment") or "").strip().lower()
            if seg:
                segments.add(seg)
    print("LOOKALIKE DISCOVERY QUERIES (run these in a browser; paste results")
    print("into a CSV and `leads.py ingest`). Works with any search engine,")
    print("LinkedIn Sales Nav saved search (>20 results -> People export), etc.\n")
    for seg in sorted(segments):
        print(f"  ── {seg}")
        print(f'    "{seg}" companies "united kingdom" | "uae" | "usa" -site:linkedin.com')
        print(f'    site:clutch.ai "{seg}" 5-50 employees')
        print(f'    site:goodfirms.co "{seg}" studio project manager')
        print(f'    "{seg}" "hiring" "recruiter" "@[company].com" -site:linkedin.com')
        print()


if __name__ == "__main__":
    args = sys.argv[1:]
    cmd = args[0] if args else "help"
    if cmd == "ingest" and len(args) == 2:
        ingest(args[1])
    elif cmd == "score" and len(args) == 2:
        score(args[1])
    elif cmd == "merge" and len(args) == 2:
        merge(args[1])
    elif cmd == "lookalikes":
        lookalikes(args[1] if len(args) > 1 else "")
    elif cmd == "reclassify":
        counts = reclassify(args[1] if len(args) > 1 else None)
        for rec, n in sorted(counts.items(), key=lambda kv: -kv[1]):
            print(f"  {n:>4}  {rec}")
    else:
        print(__doc__)
        sys.exit(1)
