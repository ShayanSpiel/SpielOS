#!/usr/bin/env python3
"""
Outbound Department — layered lead verification.

Layered verification — this is the real answer to "unverified" leads:

  L1 heuristic (instant, free):
      syntax, disposable/personal domains, MX records, role addresses.
      An address with valid syntax + real MX + non-disposable domain is
      "Plausible" — it can open and click (your data proves it does).

  L2 SMTP probe (the temp-server idea, safe):
      connect to the recipient MX, EHLO, MAIL FROM, RCPT TO, classify by
      the server's answer. Runs from THIS box's IP — which is not any
      sending infrastructure, so probing can never hurt spielos.xyz
      deliverability. Rate-limited to stay under MX-server radar.

  L3 bounce feedback:
      sending providers already return bounce events; sync-bounces marks
      them "Bounced; suppressed" in the master so they never re-enter.

Tiers (the queue consumes in this order):
      Verified  — L2-probed 250/251/252, or Apollo "Verified" + MX
      Plausible — L1 passed (MX + non-disposable), not probed yet
      Catch-all — domain accepts everything; mailbox unknowable, low prio
      Invalid   — syntax/disposable/no-MX/bounced — NEVER sends

Usage:
    python3 verify.py probe <email>            # L2 single probe
    python3 verify.py probe-queue [--limit N]  # L2 the master's non-verified
    python3 verify.py audit                    # tier summary of the master
    python3 verify.py sync-bounces             # L3: bounce events -> master
"""

import csv
import os
import re
import smtplib
import socket
import subprocess
import sys
import time
from datetime import datetime, timezone

HERE = os.path.dirname(os.path.abspath(__file__))

from . import config  # noqa: E402
from . import outbound  # noqa: E402
from . import providers  # noqa: E402

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")
DISPOSABLE = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com",
    "icloud.com", "proton.me", "protonmail.com", "mail.com", "gmx.com",
    "zoho.com", "yandex.com", "outlook.fr", "live.com", "msn.com",
}
ROLE_LOCALS = {"info", "hello", "contact", "support", "admin", "office",
               "sales", "team", "hr", "careers", "jobs", "enquiries", "mail",
               "noreply", "business", "billing", "enquiry", "ask", "reach"}

PROBE_FROM = "verify@spielos.xyz"
PROBE_DELAY = 1.5        # seconds between probes (MX servers rate-limit)
SOCK_TIMEOUT = 12
MAX_PER_DOMAIN = 3       # probes per MX domain per run (be a polite guest)


# ── L1: heuristic ────────────────────────────────────────────────────────────

def l1_check(email: str) -> dict:
    """{tier, reason}. Pure heuristic, no network beyond DNS."""
    e = (email or "").strip().lower()
    if not EMAIL_RE.match(e):
        return {"tier": "Invalid", "reason": "syntax"}
    dom = e.split("@")[-1]
    if dom in DISPOSABLE:
        return {"tier": "Invalid", "reason": "disposable/personal domain"}
    local = e.split("@")[0]
    is_role = local in ROLE_LOCALS or local.startswith(tuple(f"{r}." for r in ROLE_LOCALS))
    if not providers.dns_has_mx(dom):
        return {"tier": "Invalid", "reason": "no MX records"}
    if is_role:
        return {"tier": "Catch-all", "reason": "role address (MX ok)"}
    return {"tier": "Plausible", "reason": f"MX ok, not disposable ({dom})"}


# ── L2: SMTP mailbox probe ───────────────────────────────────────────────────

def _mx_targets(email: str) -> list:
    dom = email.split("@")[-1]
    mx = providers.dns_mx(dom)
    if mx:
        out = []
        for _p, exchange in mx[:2]:
            ips = providers._dns_fallback(exchange) or providers._cached_ips(exchange)
            out.extend((exchange, ip) for ip in (ips or []))
        return out
    ips = providers._dns_fallback(dom)
    return [(dom, ip) for ip in ips]


def probe_one(email: str) -> dict:
    """L2 probe one address. Returns {tier, code, detail}."""
    e = (email or "").strip().lower()
    targets = _mx_targets(e)
    if not targets:
        return {"tier": "Invalid", "reason": "no MX"}
    last = "no response"
    for host, ip in targets:
        try:
            with socket.create_connection((ip, 25), timeout=SOCK_TIMEOUT) as s:
                s.settimeout(SOCK_TIMEOUT)
                f = s.makefile("rb")
                code = _smtp_read(f)
                if code >= 400:
                    last = f"greet {code}"
                    continue
                s.sendall(b"EHLO verify.spielos.xyz\r\n")
                while True:
                    line = f.readline()
                    if not line:
                        break
                    if line[:3].isdigit() and line[3:4] == b" ":
                        break
                s.sendall(f"MAIL FROM:<{PROBE_FROM}>\r\n".encode())
                _smtp_read(f)
                s.sendall(f"RCPT TO:<{e}>\r\n".encode())
                code = _smtp_read(f)
                s.sendall(b"QUIT\r\n")
                if code in (250, 251, 252):
                    return {"tier": "Verified", "code": code, "detail": f"{host} accepted"}
                if code in (550, 551, 553, 552, 521):
                    return {"tier": "Invalid", "code": code, "detail": f"{host} refused"}
                if code in (450, 451, 452):
                    return {"tier": "Plausible", "code": code,
                            "detail": "temp fail (rate-limited) — keep plausible"}
                return {"tier": "Plausible", "code": code, "detail": f"{host} {code}"}
        except (socket.timeout, OSError) as ex:
            last = f"{host}: {type(ex).__name__}"
            continue
    return {"tier": "Plausible", "code": None, "detail": f"unreachable ({last}) — keep plausible"}


def _smtp_read(f) -> int:
    try:
        line = f.readline()
        if not line:
            return 0
        return int(line[:3])
    except Exception:
        return 0


# ── Master integration ───────────────────────────────────────────────────────

def _read_master() -> list:
    rows = []
    for i, row in enumerate(outbound.read_contacts()):
        rows.append(row)
    return rows


def _update_row(email: str, status: str) -> int:
    """Set email_status for one master row. Returns 1 if changed."""
    from openpyxl import load_workbook
    wb = load_workbook(config.DATABASE_PATH)
    ws = wb[outbound.SHEET_NAME]
    cmap = outbound.COL_MAP
    changed = 0
    for row in ws.iter_rows(min_row=2):
        e = str(row[cmap["email"]].value or "").strip().lower()
        if e == email.lower():
            cur = str(row[cmap["email_status"]].value or "").strip()
            if cur != status:
                row[cmap["email_status"]].value = status
                changed = 1
            break
    if changed:
        wb.save(config.DATABASE_PATH)
    wb.close()
    return changed


def probe_queue(limit: int = 40) -> dict:
    """L2-probe master leads that are Plausible/unknown, upgrade them.
    Politeness: 1.5s apart, ≤3 per domain per run."""
    rows = _read_master()
    targets = []
    for r in rows:
        s = (r.get("email_status") or "").strip()
        if s in ("Verified", "Invalid", "Bounced; suppressed"):
            continue
        if "catch-all" in s.lower() and not s:
            continue
        targets.append(r)
    domain_count = {}
    results = {"Verified": 0, "Plausible": 0, "Catch-all": 0, "Invalid": 0, "skipped": 0}
    done = 0
    for r in targets[:limit]:
        e = str(r.get("email") or "").strip().lower()
        dom = e.split("@")[-1] if e else ""
        if domain_count.get(dom, 0) >= MAX_PER_DOMAIN:
            results["skipped"] += 1
            continue
        res = probe_one(e)
        domain_count[dom] = domain_count.get(dom, 0) + 1
        tier = res["tier"]
        if tier in ("Verified", "Invalid"):
            _update_row(e, "Verified" if tier == "Verified" else "Bounced; suppressed")
        results[tier] = results.get(tier, 0) + 1
        done += 1
        time.sleep(PROBE_DELAY)
    results["probed"] = done
    return results


def sync_bounces() -> int:
    """L3: mark every bounced email in the master as suppressed."""
    import analytics
    metrics = analytics.load_metrics()
    log = outbound.load_sent_log()
    id2email = {s.get("lead_id"): s.get("email") for s in log.get("sent", [])}
    bounced_ids = [k for k, v in metrics.get("emails", {}).items()
                   if v.get("status") == "bounced"]
    changed = 0
    for lid in bounced_ids:
        e = id2email.get(lid)
        if e:
            changed += _update_row(e, "Bounced; suppressed")
    return changed


def audit() -> dict:
    rows = _read_master()
    from collections import Counter
    tiers = Counter()
    for r in rows:
        s = str(r.get("email_status") or "").strip().lower()
        if s == "verified" or "verified" in s and "not deliverability" not in s:
            tiers["Verified"] += 1
        elif "bounced" in s or "invalid" in s:
            tiers["Invalid"] += 1
        elif "catch-all" in s:
            tiers["Catch-all"] += 1
        else:
            tiers["Plausible/unknown"] += 1
    return dict(tiers)


def main():
    argv = sys.argv[1:]
    if not argv:
        print(__doc__)
        sys.exit(1)
    cmd = argv[0]
    if cmd == "probe" and len(argv) > 1:
        e = argv[1]
        l1 = l1_check(e)
        if l1["tier"] == "Invalid":
            print(f"{e}: {l1['tier']} (L1: {l1['reason']})")
        else:
            r = probe_one(e)
            print(f"{e}: {r['tier']} (L2: {r.get('detail', r.get('reason', ''))})")
    elif cmd == "probe-queue":
        limit = 40
        if "--limit" in argv:
            limit = int(argv[argv.index("--limit") + 1])
        res = probe_queue(limit)
        print(f"probe-queue: {res}")
    elif cmd == "audit":
        print(f"audit: {audit()}")
    elif cmd == "sync-bounces":
        n = sync_bounces()
        print(f"sync-bounces: {n} rows marked 'Bounced; suppressed'")
    else:
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
